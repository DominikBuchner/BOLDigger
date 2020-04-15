import openpyxl, requests_html, datetime
import PySimpleGUI as sg
from more_itertools import unique_everseen
from boldigger.boldblast_coi import slices
from bs4 import BeautifulSoup as BSoup

## funtion to scan the BOLDresutls file and return a list
## of all unique Process ID's found, will also return the workbook
## ato later use it when saving the data
## as well as what type of result were dealing with (coi or its/rbl)
def retrieve_process_ids(xlsx_path):
    ## open resultsfile and worksheet
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active

    ## get type (coi or its/rbcl) and extract process ids
    if sheet.cell(row = 1, column = 11).value == 'Process ID':
        type = 'coi'
        process_ids = [sheet.cell(row = row, column = 11).value for row in range(2, sheet.max_row + 1) if sheet.cell(row = row, column = 11).value != None and sheet.cell(row = row, column = 11).value != '']
    elif sheet.cell(row = 1, column = 11).value == 'Similarity (%)':
        type = 'its_rbcl'
        process_ids = [sheet.cell(row = row, column = 14).value for row in range(2, sheet.max_row + 1) if sheet.cell(row = row, column = 14).value != None and sheet.cell(row = row, column = 14).value != '']
    else:
        raise ValueError

    ## remove duplicates and return
    return wb, list(unique_everseen(process_ids)), type

## function to loop through the process IDs and call the BOLD API  for additional data
## will return a dict in form of {Process ID: [BOLD Record ID, BIN, Sex, Life stage, Country, Identifier, Identification method, Institution storing]}
def get_data(process_ids, processbar):
    ## slice the process ids in parts of 100 to get the maximum out of API calls
    process_ids = list(slices(process_ids, 100))

    ## open a new html session
    with requests_html.HTMLSession() as session:

        ## initialize the final dict output
        process_id_dict = {}

        for id_pack in process_ids:

            ## request the data, retry if the request failes
            while True:
                try:
                    resp = session.get('http://www.boldsystems.org/index.php/API_Public/specimen?ids=' + '|'.join(id_pack))
                except:
                    continue
                break

            ## use beautifulsoup to find all records in the response
            soup = BSoup(resp.text, 'html.parser')
            records = soup.find_all('record')

            ## loop through records to extract interesting data
            for count in range(len(records)):

                tags = ['processid', 'record_id', 'bin_uri', 'sex', 'lifestage', 'country', 'identification_provided_by', 'identification_method', 'institution_storing']

                ## search the data encapsuled by each tag
                specimen_data = [records[count].find(tag) if records[count].find(tag) != None else '' for tag in tags]

                ## extract the texts from these tags
                specimen_data = [datapoint.text if datapoint != '' else '' for datapoint in specimen_data]

                ## append a link to the specimen page to the results if user is interested in looking up even more data
                specimen_data.append('http://www.boldsystems.org/index.php/MAS_DataRetrieval_OpenSpecimen?selectedrecordid=' + specimen_data[1])

                ## put the data in a dictionary in form of processID: [listofinformation]
                process_dict = {specimen_data[0]: specimen_data[1:]}

                ## update the final output
                process_id_dict.update(process_dict)

            processbar.UpdateBar(round(100 / len(process_ids) * (process_ids.index(id_pack) + 1)))

    return process_id_dict

## function to save the additional data to the input file
def save_results(workbook, additional_data, xlsx_path, type):

    sheet = workbook.active

    ## headers to add to the resultfile
    headers = ['Record ID', 'BOLD BIN', 'Sex', 'Life stage', 'Country', 'Identifier', 'Identification Method', 'Institution storing', 'Specimen page url']

    ## first empty column depends on the type either 12 oder 15
    start = 12 if type == 'coi' else 15

    ## write the header row
    for i in range(start, start + len(headers)):
        sheet.cell(row = 1, column = i).value = headers[i - start]

    ## add the additional data to the resultfile
    for row in range(2, sheet.max_row + 1):
        for column in range(start, start + len(headers)):
            if sheet.cell(row = row, column = start - 1).value in additional_data.keys():
                sheet.cell(row = row, column = column).value = additional_data[sheet.cell(row = row, column = start - 1).value][column - start]

    workbook.save(xlsx_path)

## main function to controll the flow
def main(xlsx_path):

    ## define a layout for the new window
    layout = [
    [sg.Text('Download status', size = (15, 1), key = 'bar_des'), sg.ProgressBar(100, orientation = 'h', size = (20, 20), key = 'bar')],
    [sg.Multiline(size = (50, 10), key = 'out', autoscroll = True)]
    ]

    ## run the download loop only once. After that only run event loop
    window = sg.Window('Additional data download', layout)
    bar = window['bar']
    ran = False

    while True:

        event, values = window.read(timeout = 10)
        ## start inner loop only once
        window['out'].print("%s: Extracting process ID's." % datetime.datetime.now().strftime("%H:%M:%S"))
        window.Refresh()

        if not ran:

            ## catch any wrong file formats here to avoid crash
            try:
                wb, process_ids, type = retrieve_process_ids(xlsx_path)
            except ValueError:
                window['out'].print('Wrong file format. Close to continue.')

            ## collect the data from BOLD
            window['out'].print('%s: Downloading additional data.' % datetime.datetime.now().strftime("%H:%M:%S"))
            window.Refresh()
            additional_data = get_data(process_ids, bar)

            ## saving the data according to type
            window['out'].print('%s: Saving the results.' % datetime.datetime.now().strftime("%H:%M:%S"))
            window.Refresh()
            save_results(wb, additional_data, xlsx_path, type)

        ran = True

        ## get events (just closing) and values
        window['out'].print('%s: Done. Close to continue.' % datetime.datetime.now().strftime("%H:%M:%S"))
        window.Refresh()
        event, values = window.read()

        if event == None:
            break

    window.Close()
